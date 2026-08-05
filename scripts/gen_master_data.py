import openpyxl, io, re, os

BASE = r"D:\Users\user\Desktop\Flexora\app\assets\Production"
OUT = r"D:\Users\user\Desktop\Flexora\app\lib\core\utils\pgpl_master_import_data.dart"

# ---- STOCK: JULY 2026 (3), column EF "Closing Stock in RMT" ----
wv = openpyxl.load_workbook(os.path.join(BASE, '1. Closing Paper Stock JULY  2026 - Copy - Copy.xlsx'), data_only=True)
ws = wv['JULY 2026 (3)']

def num(v): return v if isinstance(v, (int, float)) else 0
def st(v): return str(v).strip() if v is not None else ''
def esc(x): return x.replace('\\', '\\\\').replace("'", "\\'")
def gsm_of(raw):
    m = re.match(r'\s*(\d+(?:\.\d+)?)', raw)
    return float(m.group(1)) if m else 0.0

rows, sup = [], set()
for r in range(4, ws.max_row + 1):
    comp = st(ws.cell(r, 2).value).upper()
    if not comp or comp == 'SR NO':
        continue
    closing = num(ws.cell(r, 136).value)
    if closing <= 0:
        continue
    rawgsm = st(ws.cell(r, 4).value)
    rows.append(dict(sup=comp, mat=st(ws.cell(r, 3).value).upper(),
                     gsm=gsm_of(rawgsm), rawgsm=rawgsm,
                     code=st(ws.cell(r, 5).value),
                     size=num(ws.cell(r, 6).value), rmt=closing))
    sup.add(comp)

# ---- CLIENTS: order tracker, with user-confirmed typo merges ----
wo = openpyxl.load_workbook(os.path.join(BASE, 'New Order Detail and Status Tracking 2026-2027.xlsx'), data_only=True)
raw = set()
for sh in wo.worksheets:
    col = None
    for c in range(1, 30):
        if str(sh.cell(1, c).value or '').strip().lower().startswith('client'):
            col = c
            break
    if not col:
        continue
    for r in range(2, sh.max_row + 1):
        v = sh.cell(r, col).value
        if isinstance(v, str) and v.strip():
            raw.add(v.strip().upper())

MERGE = {
    'HARBAL HILLS': 'HERBAL HILLS', 'HERBAL HILL': 'HERBAL HILLS',
    'SHUBHSHREE ARTS': 'SHUBHASHREE ARTS', 'SHUBHSHRI ART': 'SHUBHASHREE ARTS',
    'GLOBAL CORP': 'GLOBAL CROP CARE', 'GLOBAL CORP CARE': 'GLOBAL CROP CARE',
}
clients = sorted({MERGE.get(x, x) for x in raw})

L = []
A = L.append
A("// GENERATED FROM PGPL SOURCE WORKBOOKS - DO NOT HAND-EDIT.")
A("// Regenerate with scripts/gen_master_data.py after the workbooks change.")
A("//")
A("//   Stock     : '1. Closing Paper Stock JULY  2026 - Copy - Copy.xlsx'")
A("//               sheet 'JULY 2026 (3)', column EF 'Closing Stock in RMT'.")
A("//               That column is a hardcoded physical stock-take, not a")
A("//               roll-forward, and is carried as the 01-Aug-2026 opening.")
A("//   Customers : 'New Order Detail and Status Tracking 2026-2027.xlsx'")
A("//   Suppliers : COMPANY NAME column of the stock workbook.")
A("//")
A("// Names only - GST / phone / address are deliberately left blank for the")
A("// user to fill in; nothing here is invented.")
A("")
A("/// One opening-stock line exactly as counted in the source workbook.")
A("class OpeningStockRow {")
A("  const OpeningStockRow(")
A("    this.supplier,")
A("    this.material,")
A("    this.gsmMicron,")
A("    this.productCode,")
A("    this.webSizeMm,")
A("    this.rmt,")
A("    this.gsmLabel,")
A("  );")
A("")
A("  final String supplier;")
A("  final String material;")
A("  final double gsmMicron;")
A("  final String productCode;")
A("  final double webSizeMm;")
A("  final double rmt;")
A("")
A("  /// Raw GSM text from the sheet (e.g. '80/60'), kept verbatim so the")
A("  /// parsed numeric value can always be traced back to the source.")
A("  final String gsmLabel;")
A("}")
A("")
A("class PgplMasterImportData {")
A("  PgplMasterImportData._();")
A("")
A("  /// The date the physical stock count represents.")
A("  static final DateTime openingStockDate = DateTime(2026, 8, 1);")
A("")
A("  static const List<String> supplierNames = [")
for s in sorted(sup):
    A("    '%s'," % esc(s))
A("  ];")
A("")
A("  static const List<String> customerNames = [")
for c in clients:
    A("    '%s'," % esc(c))
A("  ];")
A("")
A("  static const List<OpeningStockRow> openingStock = [")
for x in rows:
    A("    OpeningStockRow('%s', '%s', %s, '%s', %s, %s, '%s')," % (
        esc(x['sup']), esc(x['mat']), repr(float(x['gsm'])),
        esc(x['code']), repr(float(x['size'])), repr(float(x['rmt'])), esc(x['rawgsm'])))
A("  ];")
A("}")

io.open(OUT, 'w', encoding='utf-8').write('\n'.join(L) + '\n')
print("suppliers:%d  customers:%d (from %d raw)  stockRows:%d  totalRMT:%.0f" % (
    len(sup), len(clients), len(raw), len(rows), sum(x['rmt'] for x in rows)))
